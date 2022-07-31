import openpyxl
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium import webdriver
driver = webdriver.Firefox(executable_path="/home/zorro/Downloads/geckodriver")
# driver.get("https://www.google.com/")
# driver.find_element_by_xpath("//input[@class='gLFyf gsfi']").send_keys('max verstappen')
# driver.execute_script("window.scrollBy(0,1000)","")
driver.find_element_by_xpath("//input[@class='gNO89b']").click()
path = "/home/zorro/Downloads/7 A attendace.xlsx"
workbook = openpyxl.load_workbook(path)
sheet = workbook.active
rows = sheet.max_row
cols = sheet.max_column
for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(row=r,column=c).value)


